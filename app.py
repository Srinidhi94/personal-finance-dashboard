import os
import json
import argparse
from datetime import datetime
from flask import Flask, request, jsonify, render_template, redirect, url_for, send_file, send_from_directory
from werkzeug.utils import secure_filename

# Import our parsers
from parsers import extract_transactions_from_file

app = Flask(__name__)
app.jinja_env.globals.update(abs=abs)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['ALLOWED_EXTENSIONS'] = {'pdf', 'csv', 'txt'}
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max

# Data storage files
TRANSACTION_DB = 'data/transactions.json'
CATEGORIES_FILE = 'data/categories.json'
ACCOUNT_CONFIG_FILE = 'data/account_config.json'

# Ensure directories exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('data', exist_ok=True)

# Initialize or load categories
if os.path.exists(CATEGORIES_FILE):
    with open(CATEGORIES_FILE, 'r') as f:
        CATEGORIES = json.load(f)
else:
    # Default categories
    CATEGORIES = {
        "Income": {
            "keywords": ["salary", "interest", "dividend", "bonus", "gift"],
            "subcategories": {
                "Salary": ["salary", "payroll", "pay"],
                "Investment Income": ["interest", "dividend", "capital gain"],
                "Gifts": ["gift", "present"],
                "Other Income": ["bonus", "refund", "cashback"]
            }
        },
        "Food": {
            "keywords": ["grocery", "restaurant", "cafe", "food delivery", "groceries"],
            "subcategories": {
                "Groceries": ["grocery", "groceries", "supermarket"],
                "Restaurants": ["restaurant", "dining", "lunch", "dinner"],
                "Food Delivery": ["delivery", "zomato", "swiggy", "doordash", "uber eats"],
                "Coffee & Snacks": ["cafe", "coffee", "bakery", "snacks"]
            }
        },
        "Transportation": {
            "keywords": ["fuel", "gas", "parking", "public transport", "car service", "uber", "lyft"],
            "subcategories": {
                "Fuel": ["fuel", "gas", "petrol", "diesel"],
                "Public Transit": ["bus", "train", "metro", "subway", "public transport"],
                "Taxi/Rideshare": ["uber", "lyft", "ola", "taxi", "cab"],
                "Parking": ["parking"],
                "Car Service": ["service", "repair", "maintenance"]
            }
        },
        "Shopping": {
            "keywords": ["clothing", "electronics", "furniture", "amazon", "online shopping"],
            "subcategories": {
                "Clothing": ["clothing", "apparel", "fashion", "dress", "shirt", "shoes"],
                "Electronics": ["electronics", "gadget", "phone", "laptop", "computer"],
                "Home Goods": ["furniture", "decor", "appliance"],
                "Online Shopping": ["amazon", "flipkart", "myntra", "online"]
            }
        },
        "Utilities": {
            "keywords": ["electricity", "water", "gas", "internet", "phone", "mobile"],
            "subcategories": {
                "Electricity": ["electricity", "power", "light"],
                "Water": ["water"],
                "Gas": ["gas"],
                "Internet": ["internet", "broadband", "wifi"],
                "Mobile": ["mobile", "phone", "cell"]
            }
        },
        "Rent": {
            "keywords": ["rent", "lease", "housing", "apartment"],
            "subcategories": {}
        },
        "Entertainment": {
            "keywords": ["movies", "concert", "subscription", "netflix", "amazon prime"],
            "subcategories": {
                "Streaming": ["netflix", "amazon prime", "hotstar", "hulu", "disney+"],
                "Movies": ["movie", "cinema", "theatre"],
                "Music": ["concert", "spotify", "music"],
                "Other": ["entertainment", "subscription"]
            }
        },
        "Miscellaneous": {
            "keywords": ["other", "misc", "unknown"],
            "subcategories": {}
        }
    }
    with open(CATEGORIES_FILE, 'w') as f:
        json.dump(CATEGORIES, f, indent=2)

# Initialize or load account configuration
if os.path.exists(ACCOUNT_CONFIG_FILE):
    with open(ACCOUNT_CONFIG_FILE, 'r') as f:
        ACCOUNT_CONFIG = json.load(f)
else:
    ACCOUNT_CONFIG = {
        "accounts": [
            {
                "bank": "HDFC",
                "account_type": "credit_card",
                "account_name": "HDFC Credit Card",
                "parser": "hdfc_credit_card"
            },
            {
                "bank": "HDFC",
                "account_type": "savings",
                "account_name": "HDFC Savings Account",
                "parser": "hdfc_savings"
            }
        ]
    }
    with open(ACCOUNT_CONFIG_FILE, 'w') as f:
        json.dump(ACCOUNT_CONFIG, f, indent=2)

# Initialize or load transactions
if os.path.exists(TRANSACTION_DB):
    with open(TRANSACTION_DB, 'r') as f:
        transactions = json.load(f)
else:
    transactions = []
    with open(TRANSACTION_DB, 'w') as f:
        json.dump(transactions, f, indent=2)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def categorize_transaction(description, amount=None, is_debit=None):
    """Categorize a transaction based on its description and amount"""
    description = (description or "").lower()
    
    # If it's a deposit (positive amount), categorize as income
    if amount is not None and amount > 0:
        return "Income"
    
    # If we know it's not a debit transaction, it's income
    if is_debit is not None and is_debit == False:
        return "Income"
    
    # Check if this is an income-related transaction based on common deposit keywords
    income_keywords = ["salary", "interest", "dividend", "deposit", "credit", "bonus", 
                     "refund", "cashback", "income", "payment received", "add fund", 
                     "add money", "credit received", "ftd", "neft cr", "imps", "salary", 
                     "interest earned"]
    
    for keyword in income_keywords:
        if keyword in description:
            return "Income"
    
    # Check for UPI transfers that are likely to be income
    if "upi" in description and any(term in description for term in ["received", "credit", "cr", "add", "fund", "deposit"]):
        return "Income"
    
    # Check if we've stored this exact transaction before
    for transaction in transactions:
        if transaction.get('description') and transaction['description'].lower() == description and transaction.get('category'):
            return transaction['category']
    
    # Rule-based categorization
    for category, category_data in CATEGORIES.items():
        keywords = category_data.get("keywords", [])
        for keyword in keywords:
            if keyword.lower() in description:
                return category
    
    # Default category
    return "Miscellaneous"

def categorize_subcategory(description, category):
    """Determine the subcategory based on description and main category"""
    description = (description or "").lower()
    
    # Check if category exists in our structure
    if category not in CATEGORIES:
        return ""
    
    # Check if we've seen this exact transaction before
    for transaction in transactions:
        if (transaction['description'].lower() == description and
            transaction['category'] == category and
            transaction.get('subcategory')):
            return transaction['subcategory']
    
    # Rule-based subcategory assignment
    subcategories = CATEGORIES[category].get("subcategories", {})
    for subcategory, keywords in subcategories.items():
        for keyword in keywords:
            if keyword.lower() in description:
                return subcategory
    
    # Default to empty subcategory
    return ""

def generate_summary_data():
    """Generate summary statistics for the dashboard"""
    if not transactions:
        return {
            'total_transactions': 0,
            'total_income': 0,
            'total_expenses': 0,
            'net_balance': 0,
            'account_summary': [],
            'category_summary': []
        }
    
    # Basic summary
    total_transactions = len(transactions)
    total_income = sum(t['amount'] for t in transactions if t['category'] == 'Income')
    total_expenses = sum(abs(t['amount']) for t in transactions if t['category'] != 'Income')
    net_balance = total_income - total_expenses
    
    # Account summary
    account_summary = []
    account_dict = {}
    
    for t in transactions:
        account_name = t['account_name']
        if account_name not in account_dict:
            account_dict[account_name] = {
                'name': account_name,
                'type': t['account_type'],
                'bank': t['bank'],
                'income': 0,
                'expenses': 0,
                'balance': 0
            }
        
        if t['category'] == 'Income':
            account_dict[account_name]['income'] += t['amount']
        else:
            account_dict[account_name]['expenses'] += abs(t['amount'])
    
    for account in account_dict.values():
        account['balance'] = account['income'] - account['expenses']
        account_summary.append(account)
    
    # Category summary
    category_summary = []
    category_dict = {}
    
    for t in transactions:
        category = t['category']
        if category not in category_dict:
            category_dict[category] = {
                'name': category,
                'total': 0,
                'count': 0,
                'month_change': 0
            }
        
        if category != 'Income':  # Only count expenses for non-income categories
            category_dict[category]['total'] += abs(t['amount'])
            category_dict[category]['count'] += 1
        else:
            # For income category, count but don't add to expense totals
            category_dict[category]['count'] += 1
    
    for category in category_dict.values():
        category_summary.append(category)
    
    # Sort by highest expense
    category_summary.sort(key=lambda x: x['total'], reverse=True)
    
    return {
        'total_transactions': total_transactions,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_balance': net_balance,
        'account_summary': account_summary,
        'category_summary': category_summary
    }

@app.route('/')
def index():
    # Generate summary data for combined dashboard & analytics
    summary_data = generate_summary_data()
    
    return render_template('index.html', 
                          transactions=transactions[-20:], 
                          categories=CATEGORIES,
                          accounts=ACCOUNT_CONFIG['accounts'],
                          summary=summary_data)

# Add a route for static files (helpful for debugging)
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(app.static_folder, filename)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return redirect(request.url)
    
    file = request.files['file']
    bank = request.form.get('bank', 'Unknown')
    account_type = request.form.get('account_type', 'Unknown')
    account_name = request.form.get('account_name', 'Unknown')
    
    if file.filename == '':
        return redirect(request.url)
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Extract transactions
        new_transactions = extract_transactions_from_file(
            filepath, bank, account_type, account_name
        )
        
        # Categorize transactions
        for transaction in new_transactions:
            amount = transaction.get('amount', 0)
            is_debit = transaction.get('is_debit', None)
            category = categorize_transaction(transaction['description'], amount, is_debit)
            subcategory = categorize_subcategory(transaction['description'], category)
            transaction['category'] = category
            transaction['subcategory'] = subcategory
        
        # Add to our database
        global transactions
        transactions.extend(new_transactions)
        
        # Save updated transactions
        with open(TRANSACTION_DB, 'w') as f:
            json.dump(transactions, f, indent=2)
        
        return redirect(url_for('review_transactions', 
                              batch_id=len(transactions)-len(new_transactions)))
    
    return redirect(url_for('index'))

@app.route('/review/<int:batch_id>')
def review_transactions(batch_id):
    # Get transactions starting from batch_id
    batch = transactions[batch_id:]
    return render_template('review.html', 
                          transactions=batch, 
                          batch_id=batch_id, 
                          categories=CATEGORIES)

@app.route('/update_transaction', methods=['POST'])
def update_transaction():
    data = request.json
    transaction_idx = data['transaction_idx']
    new_category = data['category']
    new_subcategory = data.get('subcategory', '')
    new_description = data.get('description', None)
    
    # Update category and subcategory
    transactions[transaction_idx]['category'] = new_category
    transactions[transaction_idx]['subcategory'] = new_subcategory
    
    # Update description if provided
    if new_description is not None:
        transactions[transaction_idx]['description'] = new_description
    
    # Save updated transactions
    with open(TRANSACTION_DB, 'w') as f:
        json.dump(transactions, f, indent=2)
    
    return jsonify({'success': True})

@app.route('/export', methods=['GET'])
def export_data():
    format_type = request.args.get('format', 'excel')
    
    if format_type == 'excel':
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Convert transactions to DataFrame
        df = pd.DataFrame(transactions)
        
        # Create workbook
        wb = Workbook()
        ws_trans = wb.active
        ws_trans.title = "Transactions"
        
        # Add headers
        headers = ["Date", "Description", "Amount", "Category", "Subcategory", "Account", "Bank"]
        ws_trans.append(headers)
        
        # Style headers
        for col in range(1, len(headers) + 1):
            cell = ws_trans.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9D9D9")
        
        # Add data
        for t in transactions:
            ws_trans.append([
                t['date'],
                t['description'],
                t['amount'],
                t['category'],
                t.get('subcategory', ''),
                t['account_name'],
                t['bank']
            ])
        
        # Add category summary sheet
        ws_cat = wb.create_sheet(title="Category Summary")
        
        # Add headers
        ws_cat.append(["Category", "Total Amount", "Transaction Count"])
        
        # Style headers
        for col in range(1, 4):
            cell = ws_cat.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9D9D9")
        
        # Add category data
        category_summary = {}
        for t in transactions:
            if t['amount'] < 0:  # Consider only expenses
                category = t['category']
                if category not in category_summary:
                    category_summary[category] = {'total': 0, 'count': 0}
                
                category_summary[category]['total'] += abs(t['amount'])
                category_summary[category]['count'] += 1
        
        row = 2
        for category, data in sorted(category_summary.items(), 
                                    key=lambda x: x[1]['total'], 
                                    reverse=True):
            ws_cat.append([
                category,
                data['total'],
                data['count']
            ])
            row += 1
        
        # Add account summary sheet
        ws_acc = wb.create_sheet(title="Account Summary")
        
        # Add headers
        ws_acc.append(["Account", "Income", "Expenses", "Balance"])
        
        # Style headers
        for col in range(1, 5):
            cell = ws_acc.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9D9D9")
        
        # Add account data
        account_summary = {}
        for t in transactions:
            account = t['account_name']
            if account not in account_summary:
                account_summary[account] = {'income': 0, 'expenses': 0}
            
            if t['amount'] > 0:
                account_summary[account]['income'] += t['amount']
            else:
                account_summary[account]['expenses'] += abs(t['amount'])
        
        row = 2
        for account, data in account_summary.items():
            balance = data['income'] - data['expenses']
            ws_acc.append([
                account,
                data['income'],
                data['expenses'],
                balance
            ])
            row += 1
        
        # Save the workbook
        output_file = 'expense_tracker_export.xlsx'
        wb.save(output_file)
        
        return send_file(output_file, as_attachment=True)
    
    elif format_type == 'csv':
        import pandas as pd
        
        # Convert transactions to DataFrame
        df = pd.DataFrame(transactions)
        
        # Select relevant columns
        if not df.empty:
            columns = ['date', 'description', 'amount', 'category', 'subcategory', 
                       'account_name', 'bank', 'account_type']
            export_df = df[columns]
        else:
            export_df = pd.DataFrame(columns=['date', 'description', 'amount', 'category', 
                                              'subcategory', 'account_name', 'bank', 'account_type'])
        
        # Save to CSV
        output_file = 'expense_tracker_export.csv'
        export_df.to_csv(output_file, index=False)
        
        return send_file(output_file, as_attachment=True)
    
    return jsonify({'success': False, 'message': 'Invalid format type'})

@app.route('/api/transactions', methods=['GET'])
def get_transactions():
    return jsonify(transactions)

@app.route('/api/categories', methods=['GET'])
def get_categories():
    return jsonify(CATEGORIES)

@app.route('/api/accounts', methods=['GET'])
def get_accounts():
    return jsonify(ACCOUNT_CONFIG['accounts'])

@app.route('/dashboard')
def dashboard():
    # Redirect to home page since we've merged dashboard and analytics
    return redirect(url_for('index'))

@app.route('/transactions')
def view_transactions():
    # Filter parameters
    category = request.args.get('category', '')
    account = request.args.get('account', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Filter transactions
    filtered_transactions = transactions.copy()
    
    if category:
        filtered_transactions = [t for t in filtered_transactions if t['category'] == category]
    
    if account:
        filtered_transactions = [t for t in filtered_transactions if t['account_name'] == account]
    
    if date_from:
        filtered_transactions = [t for t in filtered_transactions if t['date'] >= date_from]
    
    if date_to:
        filtered_transactions = [t for t in filtered_transactions if t['date'] <= date_to]
    
    # Sort by date (oldest first)
    # Convert date strings to datetime objects for proper sorting
    def parse_date_for_sorting(date_str):
        try:
            # Parse DD/MM/YYYY format (most common)
            return datetime.strptime(date_str, "%d/%m/%Y")
        except ValueError:
            try:
                # Try YYYY-MM-DD format
                return datetime.strptime(date_str, "%Y-%m-%d")
            except ValueError:
                try:
                    # Try DD-MM-YYYY format
                    return datetime.strptime(date_str, "%d-%m-%Y")
                except ValueError:
                    # If all else fails, return a very old date
                    # This ensures unparseable dates are at the beginning
                    # and can be easily identified and fixed
                    print(f"Warning: Could not parse date format: {date_str}")
                    return datetime(1900, 1, 1)
    
    # Sort transactions by date (oldest first)
    filtered_transactions.sort(key=lambda x: parse_date_for_sorting(x['date']))
    
    return render_template('transactions.html', 
                          transactions=filtered_transactions,
                          categories=CATEGORIES,
                          accounts=ACCOUNT_CONFIG['accounts'])

@app.route('/api/charts/category_distribution', methods=['GET'])
def chart_category_distribution():
    """API endpoint for category distribution chart"""
    if not transactions:
        return jsonify({})
    
    # Filter expenses only
    expenses = [t for t in transactions if t['amount'] < 0]
    
    # Group by category
    categories = {}
    for t in expenses:
        category = t['category']
        if category not in categories:
            categories[category] = 0
        categories[category] += abs(t['amount'])
    
    # Format for chart
    labels = list(categories.keys())
    values = list(categories.values())
    
    # Generate colors
    colors = [f'hsl({hash(label) % 360}, 70%, 60%)' for label in labels]
    
    return jsonify({
        'labels': labels,
        'datasets': [{
            'data': values,
            'backgroundColor': colors
        }]
    })

@app.route('/api/charts/account_distribution', methods=['GET'])
def chart_account_distribution():
    """API endpoint for account distribution chart"""
    if not transactions:
        return jsonify({})
    
    # Group by account
    accounts = {}
    for t in transactions:
        account = t['account_name']
        if account not in accounts:
            accounts[account] = {
                'income': 0,
                'expenses': 0
            }
        
        if t['amount'] > 0:
            accounts[account]['income'] += t['amount']
        else:
            accounts[account]['expenses'] += abs(t['amount'])
    
    # Format for chart
    labels = list(accounts.keys())
    income_data = [accounts[acc]['income'] for acc in labels]
    expense_data = [accounts[acc]['expenses'] for acc in labels]
    
    return jsonify({
        'labels': labels,
        'datasets': [
            {
                'label': 'Income',
                'data': income_data,
                'backgroundColor': 'rgba(75, 192, 192, 0.6)',
                'borderColor': 'rgba(75, 192, 192, 1)',
                'borderWidth': 1
            },
            {
                'label': 'Expenses',
                'data': expense_data,
                'backgroundColor': 'rgba(255, 99, 132, 0.6)',
                'borderColor': 'rgba(255, 99, 132, 1)',
                'borderWidth': 1
            }
        ]
    })

@app.route('/bulk_update_categories', methods=['POST'])
def bulk_update_categories():
    data = request.json
    transaction_indices = data['transaction_indices']
    new_category = data['category']
    
    try:
        updated_count = 0
        for idx in transaction_indices:
            if idx < len(transactions):
                transactions[idx]['category'] = new_category
                updated_count += 1
        
        # Save updated transactions
        with open(TRANSACTION_DB, 'w') as f:
            json.dump(transactions, f, indent=2)
        
        return jsonify({'success': True, 'updated_count': updated_count})
    except Exception as e:
        print(f"Error in bulk update: {e}")
        return jsonify({'success': False, 'error': str(e)})

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Personal Finance Dashboard')
    parser.add_argument('--port', type=int, default=5000, help='Port to run the server on')
    args = parser.parse_args()
    
    print(f"[DEBUG] Starting server on port {args.port}")
    app.run(host='0.0.0.0', port=args.port, debug=True)
